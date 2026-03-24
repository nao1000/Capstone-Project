# import os
# import sys
# import django

# # 1. Set up path first
# BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# INNER_DIR = os.path.join(BASE_DIR, 'scheduler')
# sys.path.insert(0, INNER_DIR)

# # 2. Configure Django
# os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'scheduler.settings')
# django.setup()

# from django.contrib.auth.models import User
# from scheduler.core.models import Team, Role, AvailabilityRange, TeamRoleAssignment, Schedule
# from datetime import time
# import uuid
# # --- USERS ---
# users_data = [
#     {"username": "employee1", "password": "test1234", "email": "employee1@test.com", "first_name": "John", "last_name": "Doe"},
#     {"username": "employee2", "password": "test1234", "email": "employee2@test.com", "first_name": "Jane", "last_name": "Smith"},
#     {"username": "employee3", "password": "test1234", "email": "employee3@test.com", "first_name": "Bob", "last_name": "Johnson"},
#     {"username": "supervisor1", "password": "test1234", "email": "supervisor1@test.com", "first_name": "Alice", "last_name": "Williams"},
# ]

# created_users = {}
# for u in users_data:
#     user, created = User.objects.get_or_create(username=u["username"], defaults={
#         "email": u["email"],
#         "first_name": u["first_name"],
#         "last_name": u["last_name"]
#     })
#     if created:
#         user.set_password(u["password"])
#         user.save()
#         print(f"Created user: {u['username']}")
#     else:
#         print(f"User already exists: {u['username']}")
#     created_users[u["username"]] = user

# supervisor = created_users["supervisor1"]
# employees = [created_users["employee1"], created_users["employee2"], created_users["employee3"]]

# # --- TEAM ---
# team, created = Team.objects.get_or_create(
#     name="Test Team",
#     defaults={
#         "owner": supervisor,
#         "join_code": str(uuid.uuid4())
#     }
# )
# if created:
#     print(f"Created team: {team.name}")
# else:
#     print(f"Team already exists: {team.name}")

# # Add employees to team
# for emp in employees:
#     team.members.add(emp)
# print(f"Added employees to team")

# # --- AVAILABILITY RANGES ---
# # Each employee has availability Mon-Fri with slightly different hours
# availability_data = [
#     # John Doe - available most of the week
#     {"user": employees[0], "ranges": [
#         {"day": "mon", "start": "08:00", "end": "17:00"},
#         {"day": "tue", "start": "08:00", "end": "17:00"},
#         {"day": "wed", "start": "10:00", "end": "18:00"},
#         {"day": "thu", "start": "08:00", "end": "17:00"},
#         {"day": "fri", "start": "08:00", "end": "14:00"},
#     ]},
#     # Jane Smith - afternoon availability
#     {"user": employees[1], "ranges": [
#         {"day": "mon", "start": "12:00", "end": "20:00"},
#         {"day": "tue", "start": "12:00", "end": "20:00"},
#         {"day": "wed", "start": "12:00", "end": "20:00"},
#         {"day": "thu", "start": "14:00", "end": "20:00"},
#         {"day": "fri", "start": "12:00", "end": "18:00"},
#         {"day": "sat", "start": "10:00", "end": "16:00"},
#     ]},
#     # Bob Johnson - flexible hours
#     {"user": employees[2], "ranges": [
#         {"day": "mon", "start": "09:00", "end": "18:00"},
#         {"day": "wed", "start": "09:00", "end": "18:00"},
#         {"day": "fri", "start": "09:00", "end": "18:00"},
#         {"day": "sat", "start": "08:00", "end": "14:00"},
#         {"day": "sun", "start": "10:00", "end": "16:00"},
#     ]},
# ]

# from datetime import time

# for entry in availability_data:
#     user = entry["user"]
#     # Clear existing ranges for this user/team first
#     AvailabilityRange.objects.filter(user=user, team=team).delete()
#     for r in entry["ranges"]:
#         start_parts = r["start"].split(":")
#         end_parts = r["end"].split(":")
#         AvailabilityRange.objects.create(
#             user=user,
#             team=team,
#             day=r["day"],
#             start_time=time(int(start_parts[0]), int(start_parts[1])),
#             end_time=time(int(end_parts[0]), int(end_parts[1])),
#         )
#     print(f"Created availability for {user.username}")

# # --- DEFAULT SCHEDULE ---
# schedule, created = Schedule.objects.get_or_create(
#     team=team,
#     name="Default",
#     defaults={"is_active": True}
# )
# print(f"{'Created' if created else 'Exists'} default schedule")

# print("\nDone! Summary:")
# print(f"  Team: {team.name} (join code: {team.join_code})")
# print(f"  Supervisor: {supervisor.username}")
# print(f"  Employees: {[e.username for e in employees]}")
# print(f"  Schedule: {schedule.name} (active: {schedule.is_active})")

import os
import sys
import django
from datetime import time, datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INNER_DIR = os.path.join(BASE_DIR, 'scheduler')
sys.path.insert(0, INNER_DIR)

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'scheduler.settings')
django.setup()

from django.contrib.auth.models import User
from scheduler.core.models import Team, Role, AvailabilityRange, TeamRoleAssignment, Schedule
import uuid
import openpyxl
import csv

# =============================================================================
# CONFIGURATION
# =============================================================================

SCHEDULES_DIR = os.path.join(BASE_DIR, 'testData')

# Maps filename → (first_name, last_name)
# PDFs and duplicates are intentionally excluded.
# Skipped:
#   - All .pdf files (unparseable)
#   - Navya Balaji 2026 Schedule(Sheet1).csv  (using Official xlsx instead)
#   - Michael Halley S26 Schedule.xlsx        (using (2) version)
#   - CadynCrook_S26_Schedule - Sheet1.pdf   (PDF)
FILE_NAME_MAP = {
    "NinaKosukha_S26_Schedule.xlsx":                       ("Nina",       "Kosukha"),
    "LorenDipnarain_S26_Schedule.xlsx":                    ("Loren",      "Dipnarain"),
    "Skyller Lopez SP26 Schedule.xlsx":                    ("Skyller",    "Lopez"),
    "Navya Schedule - Official.xlsx":                      ("Navya",      "Balaji"),
    "NicholasAbeyta_S26_Schedule.xlsx":                    ("Nicholas",   "Abeyta"),
    "Thida Pookchan S26 Schedule.xlsx":                    ("Thida",      "Pookchan"),
    "VIctoria Paulson_S26_Schedule.xlsx":                  ("Victoria",   "Paulson"),
    "Faith Secrest Spring Semester schedule.xlsx":         ("Faith",      "Secrest"),
    "Nahal SI Leader Schedule.xlsx":                       ("Nahal",      ""),
    "SimoneGazman_S32_SISchedule.xlsx":                    ("Simone",     "Gazman"),
    "Ngunyi F Schedule 2026.xlsx":                         ("Ngunyi",     "F"),
    "Alicia Hemingson_S26_ schedule (1).xlsx":             ("Alicia",     "Hemingson"),
    "Subi Adhikari SP 26_gsheet.xlsx":                     ("Subi",       "Adhikari"),
    "Elise Spalding S26 Schedule .xlsx":                   ("Elise",      "Spalding"),
    "EllaGordon_S26_Schedule.xlsx":                        ("Ella",       "Gordon"),
    "Adriane Quach Spring 2025.xlsx":                      ("Adriane",    "Quach"),
    "ZoeKaliski_S25_Schedule .xlsx":                       ("Zoe",        "Kaliski"),
    "Jazmyne Jackson_SPR26_Schedule.xlsx":                 ("Jazmyne",    "Jackson"),
    "Evie Schmidt Spring '26.xlsx":                        ("Evie",       "Schmidt"),
    "BriannaSuarez_S26_Schedule .xlsx":                    ("Brianna",    "Suarez"),
    "Nathans_sched_26.xlsx":                               ("Nathan",     ""),
    "KevinTran_S26_Schedule.xlsx":                         ("Kevin",      "Tran"),
    "AlexisSchwartzberg_S26_Schedule.xlsx":                ("Alexis",     "Schwartzberg"),
    "AbbyTawney_S26 Schedule (3).xlsx":                    ("Abby",       "Tawney"),
    "Alyson Barr schedule Spring 26.xlsx":                 ("Alyson",     "Barr"),
    "Erin_Riekena-SP26_Schedule.xlsx":                     ("Erin",       "Riekena"),
    "SI 26 schedule Luke Bender.xlsx":                     ("Luke",       "Bender"),
    "Abby Bates_S26_Schedule.xlsx":                        ("Abby",       "Bates"),
    "Michael Halley -Schedule Spring 2026 (2).xlsx":       ("Michael",    "Halley"),
    "EmmaRazi_S26_Schedule.xlsx":                          ("Emma",       "Razi"),
    "Rio Lee - SI Leader Template Schedule - DOWNLOAD to your computer, EDIT, + REUPLOAD - dont save over this on Box .xlsx": ("Rio", "Lee"),
    "CadynCrook_S26_Schedule.xlsx":                        ("Cadyn",      "Crook"),
    "AbbieJacewicz_S26_Schedule.xlsx":                     ("Abbie",      "Jacewicz"),
    "SepehrSalamat_S26_Schedule.xlsx":                     ("Sepehr",     "Salamat"),
    "ReeceSandquist_Spring2026_ Schedule.xlsx":            ("Reece",      "Sandquist"),
    "CadenceRegan_S26_Schedule.xlsx":                      ("Cadence",    "Regan"),
    "Alena Mary S26 Schedule.xlsx":                        ("Alena",      "Mary"),
    "Cherry Ann Macarang Spring 2026 Schedule.xlsx":       ("Cherry",     "Macarang"),
    "Grace_Coyne_S26_Schedule.xlsx":                       ("Grace",      "Coyne"),
    "MarissaContreras_S26_Schedule.xlsx":                  ("Marissa",    "Contreras"),
    "MaddieCaes_S26_Schedule.xlsx":                        ("Maddie",     "Caes"),
    "TaylorStanton_S26_Schedule.xlsx":                     ("Taylor",     "Stanton"),
    "AndrewDinh_S26_Schedule.xlsx":                        ("Andrew",     "Dinh"),
    "SI Spring 2026 Schedule-Harshini R.xlsx":             ("Harshini",   "R"),
    "Taylor Raney S26 SI Schedule.xlsm":                   ("Taylor",     "Raney"),
    "VibhaChandrasekaran_S26_Schedule.xlsx":               ("Vibha",      "Chandrasekaran"),
    "CarlosNieves_S26_Schedule.xlsx":                      ("Carlos",     "Nieves"),
    "KaylaBjork_S26_Schedule.xlsx":                        ("Kayla",      "Bjork"),
    "TessaCallister_S26_Schedule.xlsx":                    ("Tessa",      "Callister"),
}

DAY_COLUMNS = {
    0: "sun",
    1: "mon",
    2: "tue",
    3: "wed",
    4: "thu",
    5: "fri",
}


# =============================================================================
# PARSING
# =============================================================================

def is_busy(cell_value):
    """A cell counts as busy if it has real text content."""
    if cell_value is None:
        return False
    s = str(cell_value).strip()
    return s != "" and s != "\xa0"


def parse_xlsx_schedule(filepath):
    """
    Parse a schedule xlsx/xlsm and return free time blocks per day.

    Returns: dict { 'mon': [(time(8,0), time(10,0)), ...], ... }
    Each tuple is a contiguous free block where the person has no classes.
    """
    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True, data_only=True)
    except Exception:
        wb = openpyxl.load_workbook(filepath, data_only=True)

    ws = wb.active

    # Find the header row and column offsets by looking for 'Sunday'
    time_col = None
    day_col_start = None

    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        for i, cell in enumerate(row):
            if cell == "Sunday":
                day_col_start = i      # index of Sunday column
                time_col = i - 1       # time is one column to the left
                break
        if day_col_start is not None:
            break

    if day_col_start is None:
        print(f"  WARNING: Could not find header row in {os.path.basename(filepath)}")
        return {}

    # Read all time-keyed rows
    # slots[day_index] = list of booleans (True = busy) for each 30-min slot
    slots = {i: [] for i in range(6)}  # 0=Sun, 1=Mon, 2=Tue, 3=Wed, 4=Thu, 5=Fri
    slot_times = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        time_val = row[time_col] if time_col >= 0 else None
        if not isinstance(time_val, time):
            continue

        slot_times.append(time_val)
        for day_i in range(6):
            col_i = day_col_start + day_i
            cell_val = row[col_i] if col_i < len(row) else None
            slots[day_i].append(is_busy(cell_val))

    if not slot_times:
        return {}

    # Convert busy-slot lists into free time ranges
    free_blocks = {}
    for day_i, busy_list in slots.items():
        day_str = DAY_COLUMNS[day_i]
        blocks = []
        in_free_block = False
        block_start = None

        for slot_idx, busy in enumerate(busy_list):
            slot_time = slot_times[slot_idx]

            if not busy:
                if not in_free_block:
                    in_free_block = True
                    block_start = slot_time
            else:
                if in_free_block:
                    in_free_block = False
                    blocks.append((block_start, slot_time))

        # Close any open block at end of day
        if in_free_block and block_start is not None:
            last_slot = slot_times[-1]
            end_hour = last_slot.hour + (1 if last_slot.minute == 30 else 0)
            end_minute = 30 if last_slot.minute == 0 else 0
            end_t = time(end_hour, end_minute)
            blocks.append((block_start, end_t))

        # Only include days that have at least one free block
        if blocks:
            free_blocks[day_str] = blocks

    return free_blocks


# =============================================================================
# MAIN
# =============================================================================

DEFAULT_PASSWORD = "siLeader2026!"

print("=" * 60)
print("Creating supervisor...")
print("=" * 60)

supervisor, created = User.objects.get_or_create(
    username="supervisor",
    defaults={
        "email": "supervisor@siteam.com",
        "first_name": "SI",
        "last_name": "Supervisor",
    }
)
if created:
    supervisor.set_password(DEFAULT_PASSWORD)
    supervisor.save()
    print(f"  Created supervisor: supervisor / {DEFAULT_PASSWORD}")
else:
    print(f"  Supervisor already exists")

print("\nCreating team...")
team, created = Team.objects.get_or_create(
    name="SI Leaders Spring 2026",
    defaults={
        "owner": supervisor,
        "join_code": str(uuid.uuid4()),
    }
)
if created:
    print(f"  Created team: {team.name}")
    print(f"  Join code: {team.join_code}")
else:
    print(f"  Team already exists: {team.name}")

print("\nCreating schedule...")
schedule, created = Schedule.objects.get_or_create(
    team=team,
    name="Spring 2026",
    defaults={"is_active": True}
)
print(f"  {'Created' if created else 'Exists'}: Spring 2026 schedule")

print("\n" + "=" * 60)
print("Processing schedule files...")
print("=" * 60)

skipped = []
processed = []

print(SCHEDULES_DIR)
for filename, (first_name, last_name) in FILE_NAME_MAP.items():
    filepath = os.path.join(SCHEDULES_DIR, filename)
    print(filepath)
    if not os.path.exists(filepath):
        print(f"\n  MISSING FILE: {filename}")
        skipped.append(filename)
        continue

    # Build username from name
    raw = f"{first_name}{last_name}".lower().replace(" ", "")
    username = raw if raw else first_name.lower()

    print(f"\n  {first_name} {last_name} → @{username}")

    # Create or get user
    user, created = User.objects.get_or_create(
        username=username,
        defaults={
            "email": f"{username}@siteam.com",
            "first_name": first_name,
            "last_name": last_name,
        }
    )
    if created:
        user.set_password(DEFAULT_PASSWORD)
        user.save()
        print(f"    Created user")
    else:
        print(f"    User already exists")

    # Add to team
    team.members.add(user)

    # Parse schedule file
    free_blocks = parse_xlsx_schedule(filepath)

    if not free_blocks:
        print(f"    WARNING: No availability parsed from file")
        skipped.append(filename)
        continue

    # Clear old availability for this user/team and recreate
    AvailabilityRange.objects.filter(user=user, team=team).delete()

    total_blocks = 0
    for day_str, blocks in free_blocks.items():
        for start_t, end_t in blocks:
            AvailabilityRange.objects.create(
                user=user,
                team=team,
                day=day_str,
                start_time=start_t,
                end_time=end_t,
            )
            total_blocks += 1

    print(f"    Saved {total_blocks} availability blocks across {len(free_blocks)} days")
    processed.append(filename)

print("\n" + "=" * 60)
print("SUMMARY")
print("=" * 60)
print(f"  Team:       {team.name}")
print(f"  Join code:  {team.join_code}")
print(f"  Supervisor: supervisor / {DEFAULT_PASSWORD}")
print(f"  Password for all SI leaders: {DEFAULT_PASSWORD}")
print(f"  Processed:  {len(processed)} files")
if skipped:
    print(f"  Skipped:    {len(skipped)} files")
    for f in skipped:
        print(f"    - {f}")